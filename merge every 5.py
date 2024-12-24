import openpyxl

dataframe = openpyxl.load_workbook('D:\Programing_codes\Python\Science\Else\Python\Book.xlsx')
dataframe1 = dataframe.active
print(dataframe1.max_row)
print(dataframe1.max_column)
sum=0
numrows=12
for step in range (0,int((dataframe1.max_row/12))):
  firstrow = dataframe1['A'+str(((step)*numrows)+2)].value
  avg=sum/12
  sum=0
  for row in range((step*numrows+1), (step*numrows+(numrows+1))):
      
      for col in dataframe1.iter_cols(2, 2):
        if(col[row].value is not None):
          sum= int((col[row].value)) +sum
        else:
          sum= avg +sum
          print("Avrage inserted: AVG=",avg)
        # print(row,"  ",col[row].value,"   sum = ",sum)
        # for col in dataframe1.iter_cols(1, 1):
      #   print(row,"  Time: ",col[row].value,)
  # print(dataframe1['A'+str(step*(numrows)+2)].value)
  print(firstrow , "Sum  =  ",sum)
  dataframe1['E'+str(step+2)]=firstrow
  dataframe1['F'+str(step+2)]=sum
  
print("Saving file...")
dataframe.save("new file.xlsx")
print("Finished.")


# for row in range(0, dataframe1.max_row):
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value)

# df = pd.read_csv('file_name_here.csv') 
# array = [[0]*(df['Consumption'].size)]*2
# for i in range (df['Consumption'])
#   array[0][i] = df['Consumption']
# list.append(df['Time'])
# list[1][df['Consumption'].size]=df['Consumption']
# list.append(df['Consumption'])
#list.append(df['Consumption'])
#print (df)
# sum=0
# for x in range(5):
#   print(list[0][x])
#   sum=(list[1][x])+sum
# print(sum)
# print(array)